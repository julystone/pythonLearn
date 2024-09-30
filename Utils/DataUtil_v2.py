"""
封装一个读取用例的excel类
# 实现读取用例数据
# 实现写入数据的功能
"""
import time
from typing import Any, Annotated
from pydantic import BaseModel, ConfigDict

import openpyxl
from openpyxl.styles import Font


# @dataclass
# class Case:
#     pass


class Case:
    def __repr__(self):
        string = '\n In __repr__：'
        string += '\n' + repr(self.__dict__)
        return string


class ExcelFoo(BaseModel):
    sheet_list: list
    cur_sheet: Any
    max_row: int
    max_column: int

    model_config = ConfigDict(arbitrary_types_allowed=True)


class ReadYaml:
    """
    读取yaml数据
    """

    def __init__(self, file_name):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        """
        pass


class ReadExcel(object):
    """
    读取excel数据
    """

    def __init__(self, file_name, sheet_name=None, if_new_column=None):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        :param sheet_name: 表单名字  -->  str
        """

        self.file_name = file_name
        self.sheet_name = sheet_name
        self.if_new_column = if_new_column
        self.wb = openpyxl.load_workbook(self.file_name)
        self.wb: openpyxl.Workbook
        dic = {
            "sheet_list": self.sheet_lists,
            "cur_sheet": self.cur_sheet,
            "max_row": self.get_max_rc[0],
            "max_column": self.get_max_rc[1]
        }
        self.Excel = ExcelFoo(**dic)
        print(self.Excel)

    def __del__(self):
        self.wb.close()

    @property
    def sheet_lists(self):
        return self.wb.sheetnames

    @property
    def cur_sheet(self):
        return self.wb.active

    def switch_sheet(self, sheet_name):
        self.wb[sheet_name]
        self.Excel.max_row, self.Excel.max_column = self.get_max_rc

    @property
    def get_max_rc(self):
        print(self.cur_sheet)
        return self.cur_sheet.max_row, self.cur_sheet.max_column

    def delete_redundant(self):
        for row in list(self.cur_sheet.rows)[::-1]:
            if row[0].value is None:
                self.cur_sheet.delete_rows(row[0].row, 1)
        for column in list(self.cur_sheet.columns)[::-1]:
            if column[0].value is None:
                self.cur_sheet.delete_cols(column[0].column, 1)

        self.save()

    def clear_sheet(self):
        self.wb.remove(self.wb[self.sheet_name])
        self.wb.create_sheet(self.sheet_name, 0)
        self.save()

    def clear_sheet_except_title(self):
        for row in list(self.cur_sheet.rows)[:0:-1]:
            self.cur_sheet.delete_rows(row[0].row, 1)
        self.save()

    def read_data_obj(self, sheet=None):
        """
        按行读取数据，表单所有数据
        每个用例存储在一个对象中
        :return: 返回一个列表，列表中每个元素为一个用例对象
        """
        if sheet is not None:
            self.switch_sheet(sheet)
        # 按行获取数据转换成列表
        rows_data = list(self.cur_sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        titles.append("max_column")
        # 定义一个空列表用来存储所有的用例
        if self.if_new_column is not None:
            if self.if_new_column not in titles:
                self.w_data(1, self.r_max()[1] + 1, self.if_new_column)
        cases = []
        for case in rows_data[1:]:
            # 创建一个Cases类的对象，用来保存用例数据，
            case_obj = Case()
            # data用例临时存放用例数据
            data = []
            # 判断该单元格是否为字符串类型，
            for cell in case:
                data.append(cell.value)
            data.append(len(case))
            # 将该条数据放入cases中
            case_data = list(zip(titles, data))
            for i in case_data:
                if i[0] == self.if_new_column or i[0] is None:
                    continue
                setattr(case_obj, i[0], i[1])
            setattr(case_obj, 'row', case[0].row)
            cases.append(case_obj)

        return cases

    def open(self):
        self.wb = openpyxl.load_workbook(self.file_name)
        self.cur_sheet = self.wb[self.sheet_name]
        self.save()

    def close(self):
        self.Excel.work_book.close()

    def w_data(self, row, column, data):
        self.open()
        self.cur_sheet.cell(row, column, data)
        self.save()
        self.close()

    def set_column_width(self, column, width):
        self.cur_sheet.column_dimensions[column].width = width
        self.save()

    def set_font(self, row, column, Font=Font(u'宋体', size=11, bold=True, color='000000')):
        cell = chr(ord("A") + column - 1) + str(row)
        self.cur_sheet[cell].font = Font
        self.save()

    def w_data_origin(self, row, column, data):
        self.cur_sheet.cell(row, column, data)

    def save(self):
        self.wb.save(self.file_name)

    def r_max(self):
        """
        0 为 最大行数， 1 为最大列数
        :return:
        """
        return self.max_row, self.max_column

    @property
    def latest_column_char(self):
        print(chr(self.Excel.max_column + ord('A')))
        return self.Excel.max_column + ord('A')

    @property
    def latest_row_num(self):
        return str(self.Excel.max_row + 1)


if __name__ == '__main__':
    r = ReadExcel("../SearchWords/SearchWords_v2.xlsx", sheet_name="July")
    time.sleep(2)
    r.switch_sheet(sheet_name="July")
    print(r.Excel)
    r.latest_column_char
    print(r.latest_row_num)
