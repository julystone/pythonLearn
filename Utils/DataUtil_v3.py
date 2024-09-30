"""
封装一个读取用例的excel类
# 实现读取用例数据
# 实现写入数据的功能
"""
import re
import time
from typing import Any, Annotated
from pydantic import BaseModel, ConfigDict

import openpyxl
from openpyxl.styles import Font


# @dataclass
# class Case:
#     pass


class Case:
    def __repr__(self):
        string = '\n In __repr__：'
        string += '\n' + repr(self.__dict__)
        return string


class SheetFoo(BaseModel):
    sheet_list: list
    selected_sheet: Any
    max_row: int
    max_column: int

    model_config = ConfigDict(arbitrary_types_allowed=True)


class ReadExcel(object):
    """
    读取excel数据
    """

    def __init__(self, file_name, sheet_name=None, if_new_column=None):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        :param sheet_name: 表单名字  -->  str
        """

        self.file_name = file_name
        self.if_new_column = if_new_column
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(self.file_name)
        self.wb: openpyxl.Workbook
        dic = {
            "sheet_list": self.sheet_lists,
            "selected_sheet": self.selected_sheet,
            "max_row": self.selected_sheet.max_row,
            "max_column": self.selected_sheet.max_column
        }
        self.Sheet = SheetFoo(**dic)
        print(self.Sheet)

    def __del__(self):
        self.wb.close()

    @property
    def sheet_lists(self):
        return self.wb.sheetnames

    @property
    def selected_sheet(self):
        return self.wb[self.sheet_name]

    @property
    def get_max_rc(self):
        return self.Sheet.selected_sheet.max_row, self.Sheet.selected_sheet.max_column

    @property
    def latest_column_char(self):
        return chr(self.Sheet.max_column + ord('A'))

    @property
    def latest_row_num(self):
        return str(self.Sheet.max_row + 1)

    @staticmethod
    def get_column_via_char(char: str):
        return ord(char.upper()) - ord('A') + 1

    @staticmethod
    def split_letter_and_number(s):
        pattern = r'(\D+)(\d+)'
        match = re.match(pattern, s)
        if match:
            letter = match.group(1)  # 捕获组1：字母部分
            number = match.group(2)  # 捕获组2：数字部分
            return letter, number
        else:
            return None, None

    def w_data_origin(self, row, column, data):
        self.Sheet.selected_sheet.cell(row, column, data)

    def w_data_char(self, string, data):
        self.w_data_origin(*self.get_row_column(string), data)

    def switch_sheet(self, sheet_name):
        if sheet_name not in self.sheet_lists:
            return "没有该sheet"
        self.sheet_name = sheet_name
        self.Sheet.selected_sheet = self.wb[sheet_name]
        self.Sheet.max_row, self.Sheet.max_column = self.get_max_rc

    def get_row_column(self, string):
        char, row = self.split_letter_and_number(string)
        column = 0
        for index, letter in enumerate(char[::-1]):
            column += self.get_column_via_char(letter) * pow(26, index)
        return int(row), column

    def save(self):
        self.wb.save(self.file_name)

    def read_data_obj(self, sheet=None):
        """
        按行读取数据，表单所有数据
        每个用例存储在一个对象中
        :return: 返回一个列表，列表中每个元素为一个用例对象
        """
        if sheet is not None:
            self.switch_sheet(sheet)
        # 按行获取数据转换成列表
        rows_data = list(self.Sheet.selected_sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        titles.append("max_column")
        # 定义一个空列表用来存储所有的用例
        if self.if_new_column is not None:
            if self.if_new_column not in titles:
                self.w_data_origin(1, self.Sheet.max_column + 1, self.if_new_column)
        cases = []
        for case in rows_data[1:]:
            # 创建一个Cases类的对象，用来保存用例数据，
            case_obj = Case()
            # data用例临时存放用例数据
            data = []
            # 判断该单元格是否为字符串类型，
            for cell in case:
                data.append(cell.value)
            data.append(len(case))
            # 将该条数据放入cases中
            case_data = list(zip(titles, data))
            for i in case_data:
                if i[0] == self.if_new_column or i[0] is None:
                    continue
                setattr(case_obj, i[0], i[1])
            setattr(case_obj, 'row', case[0].row)
            cases.append(case_obj)

        return cases


if __name__ == '__main__':
    r = ReadExcel("../SearchWords/SearchWords_v3.xlsx", sheet_name="July")
    time.sleep(2)
    r.switch_sheet(sheet_name="Sheet1")
    print(r.Sheet)
    print(r.latest_column_char)
    print(r.latest_row_num)
    res = r.get_row_column('ZZ100')
    print(res)
    # print(r.read_data_obj())
    # r.save()
