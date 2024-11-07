# encoding: utf-8

from collections import defaultdict
from Utils import DataUtil
import os
from openpyxl import load_workbook, Workbook


def filter_files_with_rules(folder_path):
    folder_file = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if '_input_rules_' in file and '~' not in file:
                abs_path = os.path.join(root, file)
                folder_file.append(abs_path)
    print(folder_file)
    return folder_file


def sheet_name_trim(sheet_name):
    index = sheet_name.find("_input")

    if index != -1:
        truncated_s = sheet_name[:index]  # 截取到"_input"之前的部分
    else:
        # 如果没有找到"_input"，可能你想保留整个字符串或者进行其他处理
        truncated_s = sheet_name
    print(truncated_s)  # 输出结果
    return truncated_s


# 遍历要合并的Excel文件
def workbook_merge(excel_files):
    new_wb = Workbook()
    for file_path in excel_files:
        src_wb = load_workbook(file_path)
        # 遍历当前Excel文件中的所有工作表
        for sheet_name in src_wb.sheetnames:
            src_ws = src_wb[sheet_name]
            new_ws = new_wb.create_sheet(title=sheet_name_trim(sheet_name))  # 在新工作簿中创建同名的工作表
            for row in src_ws.iter_rows(values_only=True):
                new_ws.append(row)
    # 保存新的Excel文件
    new_wb.save('merged_sheets.xlsx')


def rule_2_obj(rule_path):
    rules_file = filter_files_with_rules(rule_path)
    workbook_merge(rules_file)


if __name__ == '__main__':
    rule_2_obj("./CVMfiles_pc")
