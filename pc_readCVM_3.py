# encoding: utf-8

from collections import defaultdict
from Utils import DataUtil
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def filter_files_with_rules(folder_path):
    folder_file = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if 'ecs_sgRule' in file and '~' not in file:
                abs_path = os.path.join(root, file)
                folder_file.append(abs_path)
    print(folder_file)
    return folder_file


def extract_sg_rule(input_string):
    # 找到 "sg-" 的位置
    start_index = input_string.find("sg-")
    if start_index == -1:
        return "sg- not found"

        # 找到 "_cn" 的位置
    end_index = input_string.find("_cn")
    if end_index == -1:
        return "sg- found but _cn not found"

        # 确保 "_cn" 是在 "sg-" 之后找到的
    if end_index <= start_index + 3:
        return "sg- and _cn are too close"

        # 切片字符串，从 "sg-" 开始到 "_cn" 之前
    result = input_string[start_index:end_index]
    return result

# 遍历要合并的Excel文件
def workbook_merge(excel_files):
    wb = Workbook()
    for csv_file in excel_files:
        # 读取CSV文件内容到pandas DataFrame（假设CSV文件有相同的列结构）
        df = pd.read_csv(csv_file)

        # 创建或获取一个新的工作表，名称与CSV文件名（去掉.csv后缀）相同
        sheet_name = csv_file.replace('.csv', '')
        ws = wb.create_sheet(title=extract_sg_rule(sheet_name))

        # 添加标题
        for col_idx, col_name in enumerate(df.columns, 1):  # enumerate从1开始，因为Excel行号从1开始
            ws.cell(row=1, column=col_idx, value=col_name)

            # 将DataFrame的内容写入到工作表中
        for r_idx, row in df.iterrows():
            ws.append([cell for cell in row])

            # 保存工作簿
    wb.save("merged_sheets.xlsx")

def rule_2_obj(rule_path):
    rules_file = filter_files_with_rules(rule_path)
    workbook_merge(rules_file)


if __name__ == '__main__':
    rule_2_obj("./CVMfiles_pc")
